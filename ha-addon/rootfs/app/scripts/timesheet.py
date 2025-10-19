#!/usr/bin/env python3
"""
N8N Timesheet Generator

This script generates timesheets from JSON data passed via command line argument.
Usage: python3 timesheet.py '{json_data}'

Output: /paperwork/[sunday_date]/timesheet_[week_ending].xlsx and .pdf
"""

import os
import sys
import json
import logging
import subprocess
from datetime import datetime, timedelta, date
from openpyxl import load_workbook

# -------- Logging Setup --------
LOG_FILE = "logs/n8n_timesheet.log"
os.makedirs("logs", exist_ok=True)
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger()
logger.addHandler(logging.StreamHandler())

# -------- Date Helpers --------
def get_week_end_from_date(load_date: date) -> date:
    """Get the Sunday (week end) from a given date"""
    days_to_sunday = 6 - load_date.weekday()
    return load_date + timedelta(days=days_to_sunday)

def format_date_for_cell(dt: date) -> str:
    """Format date for Excel cell (e.g., 'Monday 19/08/25')"""
    return dt.strftime("%A %d/%m/%y")

def format_folder_date(dt: date) -> str:
    """Format date for folder name (e.g., '25-08-25')"""
    return dt.strftime("%d-%m-%y")

def parse_date_string(date_str: str) -> date:
    """Parse date string in format YYYYMMDD to date object"""
    return datetime.strptime(date_str, "%Y%m%d").date()

def parse_uk_date(date_str: str) -> date:
    """Parse UK date string in format DD/MM/YY to date object"""
    return datetime.strptime(date_str, "%d/%m/%y").date()

# -------- Data Processing Functions --------
def process_json_data(json_data):
    """
    Process JSON data and extract load data and time data
    Returns: (loads_by_day, time_data_by_day, week_end_date)
    """
    try:
        # Extract the data array from the JSON structure
        if isinstance(json_data, list) and len(json_data) > 0:
            data = json_data[0].get("data", [])
        else:
            data = json_data.get("data", [])

        if not data:
            logger.error("No data found in JSON")
            return None, None, None

        # Separate load entries from time entries
        loads = [item for item in data if "dwjdate" in item and "dwjload" in item]
        time_entries = [item for item in data if "date" in item and "start_time" in item]

        logger.info(f"Found {len(loads)} load entries and {len(time_entries)} time entries")

        # PRIORITY FIX: Determine week ending from LOAD dates first, not time entries
        week_end_date = None
        if loads:
            # Use the latest load date to determine the week ending
            load_dates = [parse_date_string(load["dwjdate"]) for load in loads]
            latest_load_date = max(load_dates)
            week_end_date = get_week_end_from_date(latest_load_date)
            logger.info(f"Week ending determined from loads: {week_end_date}")

        # Group loads by day - one entry per load (not per collection/delivery)
        loads_by_day = {}
        processed_loads = set()  # Track processed loads to avoid duplicates

        for load in loads:
            load_date = parse_date_string(load["dwjdate"])
            day_name = load_date.strftime("%A")
            load_number = load.get("dwjload")

            # Create unique key for this load
            load_key = f"{load_date}_{load_number}"

            if load_key not in processed_loads:
                processed_loads.add(load_key)

                # Find matching collection and delivery for this load
                collection = next((l for l in loads if l.get("dwjload") == load_number and l.get("dwjtype") == "C"), None)
                delivery = next((l for l in loads if l.get("dwjload") == load_number and l.get("dwjtype") == "D"), None)

                if collection and delivery:
                    if day_name not in loads_by_day:
                        loads_by_day[day_name] = []

                    loads_by_day[day_name].append({
                        "date": load_date,
                        "load_number": load_number,
                        "contractor": collection.get("dwjcust", "").upper(),
                        "car_count": collection.get("dwjvehs", 0),
                        "collection": collection.get("dwjtown", "").upper(),
                        "destination": delivery.get("dwjtown", "").upper()
                    })

        # Process time entries - but don't let them override the week ending
        time_data_by_day = {}

        for time_entry in time_entries:
            try:
                entry_date = parse_uk_date(time_entry.get("date", ""))
                day_name = time_entry.get("day", "").capitalize()

                # Only use time entry date for week ending if no loads exist
                if week_end_date is None:
                    week_end_date = get_week_end_from_date(entry_date)
                    logger.info(f"Week ending determined from time entries: {week_end_date}")

                time_data_by_day[day_name] = {
                    "start_time": time_entry.get("start_time", ""),
                    "finish_time": time_entry.get("finsh_time", ""),  # Note: typo in original data
                    "total_hours": time_entry.get("total_hours", ""),
                    "driver": time_entry.get("driver", ""),
                    "fleet_reg": time_entry.get("fleet_reg", "")
                }
            except Exception as e:
                logger.error(f"Error processing time entry: {e}")
                continue

        if week_end_date is None:
            logger.error("Could not determine week ending date from any data")
            return None, None, None

        logger.info(f"Processed loads by day: {list(loads_by_day.keys())}")
        logger.info(f"Processed time data by day: {list(time_data_by_day.keys())}")
        logger.info(f"Final week ending: {week_end_date}")

        return loads_by_day, time_data_by_day, week_end_date

    except Exception as e:
        logger.error(f"Error processing JSON data: {e}")
        return None, None, None

# -------- Main Generation Function --------
def generate_timesheet(json_data) -> tuple:
    """
    Generate timesheet from JSON data
    Returns: (excel_path, pdf_path) or (None, None) if failed
    """
    logger.info("Starting timesheet generation from JSON data")

    # Process JSON data
    loads_by_day, time_data_by_day, week_end_date = process_json_data(json_data)
    if not week_end_date:
        logger.error("Failed to process JSON data")
        return None, None

    # Determine output folder based on week ending date (Sunday)
    folder_date = format_folder_date(week_end_date)
    output_folder = os.path.join("/app/paperwork", folder_date)
    os.makedirs(output_folder, exist_ok=True)

    # Create filename: timesheet_weekending.xlsx
    filename = f"timesheet_{week_end_date.strftime('%Y-%m-%d')}"
    excel_output_path = os.path.join(output_folder, f"{filename}.xlsx")
    pdf_output_path = os.path.join(output_folder, f"{filename}.pdf")

    # Load template
    template_path = "/app/templates/timesheet.xlsx"
    try:
        wb = load_workbook(template_path)
        logger.info(f"Loaded template from {template_path}")
    except Exception as e:
        logger.error(f"Error loading template: {e}")
        return None, None

    ws = wb.active

    # Set week ending date
    try:
        ws["E5"] = format_date_for_cell(week_end_date)
    except Exception as e:
        logger.warning(f"Could not set week ending date at E5: {e}")

    # Define day row mappings (based on template structure)
    row_mapping = {
        "Monday": 8, "Tuesday": 11, "Wednesday": 14, "Thursday": 17,
        "Friday": 20, "Saturday": 23, "Sunday": 26
    }

    # Process loads for each day
    total_weekly_hours = 0

    for day, base_row in row_mapping.items():
        day_loads = loads_by_day.get(day, [])
        day_time = time_data_by_day.get(day, {})

        # Fill in load data (up to 3 loads per day, overflow goes to row 29+)
        for i, load in enumerate(day_loads[:3]):  # Limit to first 3 loads
            row_num = base_row + i

            try:
                ws.cell(row=row_num, column=3, value=load["contractor"])  # Customer (C)
            except Exception as e:
                logger.warning(f"Could not set contractor at row {row_num}, col 3: {e}")

            try:
                ws.cell(row=row_num, column=4, value=load["car_count"])   # No (D)
            except Exception as e:
                logger.warning(f"Could not set car count at row {row_num}, col 4: {e}")

            try:
                ws.cell(row=row_num, column=5, value=load["collection"])  # From (E)
            except Exception as e:
                logger.warning(f"Could not set collection at row {row_num}, col 5: {e}")

            try:
                ws.cell(row=row_num, column=6, value=load["destination"]) # To (F)
            except Exception as e:
                logger.warning(f"Could not set destination at row {row_num}, col 6: {e}")

            logger.debug(f"{day}: {load['contractor']} | {load['car_count']} cars | {load['collection']} → {load['destination']}")

        # Handle overflow loads (more than 3 per day)
        if len(day_loads) > 3:
            for i, load in enumerate(day_loads[3:], start=29):  # Start from row 29
                try:
                    ws.cell(row=i, column=2, value=format_date_for_cell(load["date"]))  # Date (B)
                    ws.cell(row=i, column=3, value=load["contractor"])  # Customer (C)
                    ws.cell(row=i, column=4, value=load["car_count"])   # No (D)
                    ws.cell(row=i, column=5, value=load["collection"])  # From (E)
                    ws.cell(row=i, column=6, value=load["destination"]) # To (F)
                except Exception as e:
                    logger.warning(f"Could not set overflow load at row {i}: {e}")

        # Fill in time data
        start_time = day_time.get("start_time", "")
        finish_time = day_time.get("finish_time", "")
        total_hours = day_time.get("total_hours", "")

        # Handle special cases like "sick"
        if start_time.lower() == "sick" or finish_time.lower() == "sick":
            start_time = "SICK"
            finish_time = "SICK"
            total_hours = "0"

        try:
            ws.cell(row=base_row, column=8, value=start_time.upper())   # Start (H)
        except Exception as e:
            logger.warning(f"Could not set start time at row {base_row}, col 8: {e}")

        try:
            ws.cell(row=base_row, column=9, value=finish_time.upper())  # Finish (I)
        except Exception as e:
            logger.warning(f"Could not set finish time at row {base_row}, col 9: {e}")

        try:
            ws.cell(row=base_row, column=10, value=total_hours.upper()) # Total (J)
        except Exception as e:
            logger.warning(f"Could not set total hours at row {base_row}, col 10: {e}")

        # Add to weekly total (if numeric)
        try:
            if total_hours and total_hours.lower() not in ["sick", "nan"]:
                total_weekly_hours += float(total_hours)
        except (ValueError, TypeError):
            pass

        logger.debug(f"{day} hours: {start_time} - {finish_time} = {total_hours}")

    # Set total weekly hours
    try:
        ws.cell(row=29, column=10, value=str(total_weekly_hours) if total_weekly_hours > 0 else "")
    except Exception as e:
        logger.warning(f"Could not set total weekly hours at row 29, col 10: {e}")

    # Set driver info and handle mileage/fleet registration
    if time_data_by_day:
        first_time_entry = next(iter(time_data_by_day.values()))
        driver = first_time_entry.get("driver", "")

        # Set driver name - handle merged cells better
        if driver:
            try:
                # Try to unmerge the cell first, then set value
                if hasattr(ws["H2"], 'coordinate') and ws["H2"].coordinate in ws.merged_cells.ranges:
                    # Find and unmerge the range containing H2
                    for merge_range in list(ws.merged_cells.ranges):
                        if ws["H2"].coordinate in merge_range:
                            ws.unmerge_cells(str(merge_range))
                            break
                ws["H2"] = driver.upper()
            except Exception as e:
                logger.warning(f"Could not set driver name at H2: {e}")
                # Try alternative cells if H2 fails
                try:
                    ws["G2"] = driver.upper()
                except Exception as e2:
                    logger.warning(f"Could not set driver name at G2 either: {e2}")

        # Collect all fleet registrations and mileage data
        fleet_regs = set()
        start_mileage = None
        end_mileage = None

        # Get mileage data from first and last working days
        working_days = [day for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                       if day in time_data_by_day and time_data_by_day[day].get("start_time")]

        for day_name, time_data in time_data_by_day.items():
            reg = time_data.get("fleet_reg", "").strip()
            if reg:
                fleet_regs.add(reg.upper())

            # Get mileage from first working day
            if day_name == working_days[0] if working_days else None:
                start_mileage = time_data.get("start_mileage", "").strip()

            # Get mileage from last working day
            if day_name == working_days[-1] if working_days else None:
                end_mileage = time_data.get("end_mileage", "").strip()

        # Set start mileage (H4)
        try:
            ws["H4"] = start_mileage if start_mileage else "0"
        except Exception as e:
            logger.warning(f"Could not set start mileage at H4: {e}")

        # Set end mileage (H5)
        try:
            ws["H5"] = end_mileage if end_mileage else "0"
        except Exception as e:
            logger.warning(f"Could not set end mileage at H5: {e}")

        # Set fleet registration (K5) - combine multiple regs with comma
        if fleet_regs:
            try:
                fleet_reg_str = ", ".join(sorted(fleet_regs))
                ws["K5"] = fleet_reg_str
            except Exception as e:
                logger.warning(f"Could not set fleet reg at K5: {e}")
    else:
        # Set default mileage values if no time data
        try:
            ws["H4"] = "0"  # Start mileage
            ws["H5"] = "0"  # End mileage
        except Exception as e:
            logger.warning(f"Could not set default mileage values: {e}")

    # Save Excel file
    try:
        wb.save(excel_output_path)
        logger.info(f"Timesheet Excel generated: {excel_output_path}")
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        return None, None

    # Return Excel path only (PDF generation disabled)
    logger.info("Timesheet generated successfully (XLSX only)")
    return excel_output_path, None

def main():
    """Main function to handle command line execution"""
    if len(sys.argv) < 2:
        logger.error("Usage: python3 timesheet.py '{json_data}'")
        print(json.dumps({"error": "No JSON data provided", "success": False}))
        sys.exit(1)

    try:
        # Parse JSON argument
        json_arg = sys.argv[1]
        json_data = json.loads(json_arg)

        # Generate timesheet
        excel_path, pdf_path = generate_timesheet(json_data)

        if excel_path:
            result = {
                "success": True,
                "excel_path": excel_path,
                "pdf_path": pdf_path,
                "message": "Timesheet generated successfully"
            }
        else:
            result = {
                "success": False,
                "error": "Failed to generate timesheet"
            }

        # Output result as JSON
        print(json.dumps(result))

    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON data: {e}")
        print(json.dumps({"error": f"Invalid JSON data: {e}", "success": False}))
        sys.exit(1)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        print(json.dumps({"error": f"Unexpected error: {e}", "success": False}))
        sys.exit(1)

if __name__ == "__main__":
    main()
