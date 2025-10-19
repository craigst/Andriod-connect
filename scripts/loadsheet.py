#!/usr/bin/env python3
"""
N8N Loadsheet Generator

This script generates loadsheets from JSON data passed via command line argument.
Usage: python3 loadsheet.py '{json_data}'

Output: /paperwork/[sunday_date]/[loadnumber]_[collection_name].xlsx and .pdf
"""

import os
import sys
import json
import random
import logging
import subprocess
from datetime import datetime, timedelta, date
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

# -------- Logging Setup --------
LOG_FILE = "logs/n8n_loadsheet.log"
os.makedirs("logs", exist_ok=True)
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger()
logger.addHandler(logging.StreamHandler())

# -------- Date Helpers --------
def get_week_end_from_date(load_date: date) -> date:
    """Get the Sunday (week end) from a given date"""
    days_to_sunday = 6 - load_date.weekday()
    return load_date + timedelta(days=days_to_sunday)

def format_date_for_cell(dt: date) -> str:
    """Format date for Excel cell (e.g., 'Monday 19/08/25')"""
    return dt.strftime("%A %d/%m/%y")

def format_folder_date(dt: date) -> str:
    """Format date for folder name (e.g., '25-08-25')"""
    return dt.strftime("%d-%m-%y")

def parse_date_string(date_str: str) -> date:
    """Parse date string in format YYYYMMDD to date object"""
    return datetime.strptime(date_str, "%Y%m%d").date()

# -------- Data Processing Functions --------
def process_json_data(json_data):
    """
    Process JSON data and extract load info and car details
    Returns: (load_info, cars_list)
    """
    try:
        # Extract the data array from the JSON structure
        if isinstance(json_data, list) and len(json_data) > 0:
            data = json_data[0].get("data", [])
        else:
            data = json_data.get("data", [])
        
        if not data:
            logger.error("No data found in JSON")
            return None, []
        
        # Separate job entries from car entries
        jobs = [item for item in data if "dwjtype" in item]
        cars = [item for item in data if "dwvvehref" in item]
        
        # Find delivery job to get main date
        # Support handover loads (type "S") as delivery
        delivery_job = next((job for job in jobs if job["dwjtype"] == "D"), None)
        handover_job = next((job for job in jobs if job["dwjtype"] == "S"), None)
        collection_job = next((job for job in jobs if job["dwjtype"] == "C"), None)
        
        # Use handover job as delivery if no delivery exists
        if handover_job and not delivery_job:
            delivery_job = handover_job.copy()
            # Override destination for handover loads
            delivery_job["dwjname"] = "BTT YARD FOR HAND OVER"
            delivery_job["dwjpostco"] = ""
            delivery_job["dwjtown"] = "BTT YARD FOR HAND OVER"
            logger.info("Handover load detected - using BTT YARD FOR HAND OVER as destination")
        
        if not delivery_job or not collection_job:
            logger.error("Missing delivery or collection job data")
            return None, []
        
        # Build location mapping for multi-location loads
        # Key: dwjcust + dwjadrcod, Value: location name + postcode
        collection_locations = {}
        delivery_locations = {}
        
        for job in jobs:
            location_key = f"{job.get('dwjcust', '')}_{job.get('dwjadrcod', '')}"
            location_name = f"{job.get('dwjname', '')} - {job.get('dwjpostco', '')}"
            
            if job["dwjtype"] == "C":
                collection_locations[location_key] = location_name
            elif job["dwjtype"] == "D":
                delivery_locations[location_key] = location_name
        
        # Determine if this is a multi-location load
        is_multi_collection = len(collection_locations) > 1
        is_multi_delivery = len(delivery_locations) > 1
        
        logger.info(f"Multi-location load: Collections={is_multi_collection} ({len(collection_locations)}), Deliveries={is_multi_delivery} ({len(delivery_locations)})")
        
        # Extract load info
        load_date = parse_date_string(delivery_job["dwjdate"])
        load_info = {
            "load_date": load_date,
            "load_number": delivery_job["dwjload"],
            "collections": f"{collection_job['dwjname']} - {collection_job['dwjpostco']}",
            "deliveries": f"{delivery_job['dwjname']} - {delivery_job['dwjpostco']}",
            "dwjcust": delivery_job["dwjcust"],
            "collection_name": collection_job["dwjname"].replace(" ", "_").replace("/", "_")
        }
        
        # Process cars data
        processed_cars = []
        for car in cars:
            # Apply business logic - handle null values properly
            offload_val = car.get("offload")
            offloaded = (offload_val or "N").upper()
            
            docs_val = car.get("docs")
            # If docs is null/empty, default to "Y", but if offloaded is "Y" then set to "N"
            docs = (docs_val or "Y").upper() if offloaded != "Y" else "N"
            
            spare_keys_val = car.get("sparekeys")
            
            # If spare_keys is null, set to Y, if offloaded is Y set to N
            if spare_keys_val is None:
                spare_keys = "Y" if offloaded != "Y" else "N"
            else:
                spare_keys = spare_keys_val.upper() if offloaded != "Y" else "N"
            
            # Get car notes
            car_notes = car.get("car_notes") or ""
            
            # Add location info for multi-location loads
            location_info = []
            
            if is_multi_collection:
                col_key = f"{car.get('dwvcolcus', '')}_{car.get('dwvcolcod', '')}"
                if col_key in collection_locations:
                    location_info.append(f"FROM: {collection_locations[col_key]}")
            
            if is_multi_delivery:
                del_key = f"{car.get('dwvdelcus', '')}_{car.get('dwvdelcod', '')}"
                if del_key in delivery_locations:
                    location_info.append(f"TO: {delivery_locations[del_key]}")
            
            # Combine notes with location info
            if location_info:
                location_text = " | ".join(location_info)
                final_notes = f"{location_text}\n{car_notes}".strip() if car_notes else location_text
            else:
                final_notes = car_notes
            
            processed_car = {
                "reg": car.get("dwvvehref", ""),
                "make_model": car.get("dwvmoddes") or "N/A",
                "offloaded": offloaded,
                "docs": docs,
                "spare_keys": spare_keys,
                "notes": final_notes
            }
            processed_cars.append(processed_car)
        
        logger.info(f"Processed {len(processed_cars)} cars: {[car['reg'] for car in processed_cars]}")
        
        return load_info, processed_cars
        
    except Exception as e:
        logger.error(f"Error processing JSON data: {e}")
        return None, []

def generate_load_summary(cars: list) -> str:
    """Generate load summary text"""
    not_offloaded = sum(1 for car in cars if car.get("offloaded", "").upper() != "Y")
    with_docs = sum(1 for car in cars if car.get("docs", "").upper() == "Y")
    with_spare = sum(1 for car in cars if car.get("spare_keys", "").upper() == "Y")
    
    summary = f"{not_offloaded} CARS LOADED"
    if with_docs > 0:
        summary += f", {with_docs} CARS HAVE DOCUMENTS AND HAVE BEEN PLACED ON THE PASSENGER SEAT"
    else:
        summary += ", 0 CARS HAVE DOCUMENTS"
    summary += f", {with_spare} CARS HAVE SPARE KEYS"
    return summary

def add_signature(sig_folder, cell, sheet, custom_sig_path=None):
    """Add a signature image to the given Excel sheet cell"""
    sig_path = None
    
    if custom_sig_path and os.path.exists(custom_sig_path):
        sig_path = custom_sig_path
    elif os.path.isdir(sig_folder):
        sig_files = [f for f in os.listdir(sig_folder) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
        if sig_files:
            sig_path = os.path.join(sig_folder, random.choice(sig_files))
    
    if sig_path:
        logger.info(f"Adding signature from {sig_path} to {cell}")
        img = OpenpyxlImage(sig_path)
        
        # Apply default scaling
        img.width = int(img.width * 1.0)
        img.height = int(img.height * 1.0)
        
        img.anchor = cell
        sheet.add_image(img)
        logger.info(f"✔ Signature Added: {os.path.basename(sig_path)} at {cell}")
    else:
        logger.warning(f"No signature files found in {sig_folder}")

# -------- Main Generation Function --------
def generate_loadsheet(json_data, signature_path=None) -> tuple:
    """
    Generate loadsheet from JSON data
    Returns: (excel_path, pdf_path) or (None, None) if failed
    """
    logger.info("Starting loadsheet generation from JSON data")
    
    # Process JSON data
    load_info, cars = process_json_data(json_data)
    if not load_info:
        logger.error("Failed to process JSON data")
        return None, None
    
    # Generate load summary (only the summary, not individual car notes)
    load_summary = generate_load_summary(cars)
    overall_note = load_summary
    
    # Determine output folder based on week ending date (Sunday)
    week_end_date = get_week_end_from_date(load_info["load_date"])
    folder_date = format_folder_date(week_end_date)
    output_folder = os.path.join("/app/paperwork", folder_date)
    os.makedirs(output_folder, exist_ok=True)
    
    # Create filename: loadnumber_collection.xlsx
    filename = f"{load_info['load_number']}_{load_info['collection_name']}"
    excel_output_path = os.path.join(output_folder, f"{filename}.xlsx")
    pdf_output_path = os.path.join(output_folder, f"{filename}.pdf")
    
    # Load template
    template_path = "/app/templates/loadsheet.xlsx"
    try:
        wb = load_workbook(template_path)
        logger.info(f"Loaded template from {template_path}")
    except Exception as e:
        logger.error(f"Error loading template: {e}")
        return None, None
    
    ws = wb.active
    
    # Map fixed cells
    ws["C6"] = format_date_for_cell(load_info["load_date"])  # Date
    ws["G6"] = load_info["load_number"]  # Load number
    ws["B9"] = load_info["collections"].upper()  # Collection point
    ws["F9"] = load_info["deliveries"].upper()  # Delivery point
    ws["C46"] = format_date_for_cell(load_info["load_date"])  # Collection signature date
    ws["H46"] = format_date_for_cell(load_info["load_date"])  # Delivery signature date
    
    # Map car details (up to 8 cars)
    car_cell_map = {
        0: {"make_model": "B11", "reg": "B13", "offloaded": "E10", "docs": "G10", "spare_keys": "I10", "notes": "C11"},
        1: {"make_model": "B15", "reg": "B17", "offloaded": "E14", "docs": "G14", "spare_keys": "I14", "notes": "C15"},
        2: {"make_model": "B19", "reg": "B21", "offloaded": "E18", "docs": "G18", "spare_keys": "I18", "notes": "C19"},
        3: {"make_model": "B23", "reg": "B25", "offloaded": "E22", "docs": "G22", "spare_keys": "I22", "notes": "C23"},
        4: {"make_model": "B27", "reg": "B29", "offloaded": "E26", "docs": "G26", "spare_keys": "I26", "notes": "C27"},
        5: {"make_model": "B31", "reg": "B33", "offloaded": "E30", "docs": "G30", "spare_keys": "I30", "notes": "C31"},
        6: {"make_model": "B35", "reg": "B37", "offloaded": "E34", "docs": "G34", "spare_keys": "I34", "notes": "C35"},
        7: {"make_model": "B39", "reg": "B41", "offloaded": "E38", "docs": "G38", "spare_keys": "I38", "notes": "C39"}
    }
    
    for idx in range(8):
        mapping = car_cell_map.get(idx)
        if idx < len(cars):
            car = cars[idx]
            ws[mapping["make_model"]] = car["make_model"].upper()
            ws[mapping["reg"]] = car["reg"].upper()
            ws[mapping["offloaded"]] = car["offloaded"].upper()
            ws[mapping["docs"]] = car["docs"].upper()
            ws[mapping["spare_keys"]] = car["spare_keys"].upper()
            ws[mapping["notes"]] = car["notes"].upper()
        else:
            # Clear unused car slots
            for field in mapping.values():
                ws[field] = ""
    
    # Add load summary
    ws["C39"] = overall_note.upper()
    
    # Add signatures
    try:
        # Collection signature (sig1)
        add_signature("/app/signature/sig1", "C42", ws, signature_path)
        
        # Delivery signature (sig2)
        add_signature("/app/signature/sig2", "H42", ws, signature_path)
    except Exception as e:
        logger.error(f"Error adding signatures: {e}")
    
    # Save Excel file
    try:
        wb.save(excel_output_path)
        logger.info(f"Loadsheet Excel generated: {excel_output_path}")
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        return None, None
    
    # Convert to PDF using LibreOffice
    try:
        logger.info("Converting Excel to PDF using LibreOffice...")
        
        # Run LibreOffice headless conversion
        conversion_result = subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--convert-to", "pdf",
                "--outdir", output_folder,
                excel_output_path
            ],
            capture_output=True,
            text=True,
            timeout=60
        )
        
        if conversion_result.returncode == 0 and os.path.exists(pdf_output_path):
            logger.info(f"PDF generated: {pdf_output_path}")
        else:
            logger.warning(f"PDF conversion failed: {conversion_result.stderr}")
            pdf_output_path = None
    except Exception as e:
        logger.error(f"Error converting to PDF: {e}")
        pdf_output_path = None
    
    # Return both Excel and PDF paths
    logger.info("Loadsheet generated successfully")
    return excel_output_path, pdf_output_path

def main():
    """Main function to handle command line execution"""
    if len(sys.argv) < 2:
        logger.error("Usage: python3 loadsheet.py '{json_data}'")
        print(json.dumps({"error": "No JSON data provided", "success": False}))
        sys.exit(1)
    
    try:
        # Parse JSON argument
        json_arg = sys.argv[1]
        json_data = json.loads(json_arg)
        
        # Optional signature path
        signature_path = sys.argv[2] if len(sys.argv) > 2 else None
        
        # Generate loadsheet
        excel_path, pdf_path = generate_loadsheet(json_data, signature_path)
        
        if excel_path:
            result = {
                "success": True,
                "excel_path": excel_path,
                "pdf_path": pdf_path,
                "message": "Loadsheet generated successfully"
            }
        else:
            result = {
                "success": False,
                "error": "Failed to generate loadsheet"
            }
        
        # Output result as JSON
        print(json.dumps(result))
        
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON data: {e}")
        print(json.dumps({"error": f"Invalid JSON data: {e}", "success": False}))
        sys.exit(1)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        print(json.dumps({"error": f"Unexpected error: {e}", "success": False}))
        sys.exit(1)

if __name__ == "__main__":
    main()
